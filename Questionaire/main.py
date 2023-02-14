# from tkinter.ttk import Checkbutton
import numpy as np
from openpyxl import Workbook, load_workbook


def questionnaire():
    social_style = ['Analytical', 'Amiable', 'Expressive', 'Driving']

    answers = [
        ('a', 'b', 'c', 'd'),
        ('b', 'a', 'c', 'd'),
        ('b', 'c', 'd', 'a'),
        ('d', 'a', 'c', 'b')
    ]

    counter = np.zeros((len(social_style),), dtype=np.int)
    qualities = [''] * len(answers)

    print('Enter the response:')
    for i in range(len(answers)):
        while True:
            print(f'Answer {i}: ', end='')
            option = input()
            if option in answers[i]:
                break
            print('Please enter the correct option')

        index = answers[i].index(option)
        counter[index] += 1
        qualities[i] = social_style[index]

    # prepare sheet data
    column_header = [''] + social_style
    row_header = [''] * (len(answers) + 2)
    sheet_data = np.zeros((len(answers) + 2, len(social_style)), dtype=np.int)

    row = 0
    while row < len(answers):
        column = social_style.index(qualities[row])
        sheet_data[row][column] = 1
        row_header[row] = f'Question {row + 1}'
        row += 1

    sheet_data[row] = counter
    row_header[row] = 'Total'
    row += 1

    counter *= 5

    sheet_data[row] = counter
    row_header[row] = 'Total x 5'
    row += 1

    print('Result:')
    for i in range(len(social_style)):
        print(f'{social_style[i]} : {counter[i]}%')

    return sheet_data, column_header, row_header


def generate_sheet(data, column_header, row_header):
    workbook_name = input('Enter workbook name: ')
    if not workbook_name.endswith('.xlsx'):
        workbook_name += '.xlsx'
    worksheet_title = input('Enter worksheet title: ')
    # try:
    #     workbook = load_workbook(workbook_name)
    #     if worksheet_title in workbook.sheetnames:
    #         print('Sheet already exist')
    #         choice = input('Do you want to override? Y/N')
    #         if choice.upper() == 'Y':
    #             workbook.remove(workbook[worksheet_title])
    #         else:
    #             raise InterruptedError('Operation is cancelled')
    #     worksheet = workbook.create_sheet(worksheet_title)
    #
    # except FileNotFoundError:
    #     workbook = Workbook()
    #     worksheet = workbook.active
    #     worksheet.title = worksheet_title

    # write column header
    workbook = Workbook()
    worksheet = workbook.create_sheet(worksheet_title)

    workbook.save(workbook_name)


def main():
    data, column_header, row_header = questionnaire()
    generate_sheet(data, column_header, row_header)


if __name__ == '__main__':
    main()
